from flask import Flask, request, jsonify
import base64, io, os, datetime, openpyxl, zipfile, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── palette ───────────────────────────────────────────────────────────────────
DARK_BLUE = "1F3864"; MID_BLUE  = "2E5FA3"; LIGHT_BLUE = "BDD7EE"
PALE_BLUE = "DEEAF1"; GOLD      = "FFF2CC"; GRN_HDR    = "375623"
LIGHT_GRN = "E2EFDA"; WHITE     = "FFFFFF"; LIGHT_GREY = "F2F2F2"
DARK_GREY  = "595959"; NAVY       = "1F3864"
BRIGHT_GRN = "00B050"   # GP Delta positive — bright green
OPS_DAILY  = "FCE4D6"   # salmon-orange — daily ops
OPS_WEEKLY = "DDEBF7"   # light blue    — weekly ops
OPS_MONTHLY= "E2EFDA"   # light green   — monthly ops
INSIGHT_BG = "7030A0"   # purple        — opportunity insight header

def fill(h): return PatternFill("solid", fgColor=h)
def hf(sz=10, bold=True, color=WHITE): return Font(name="Arial", size=sz, bold=bold, color=color)
def bf(sz=9, bold=False, color="000000"): return Font(name="Arial", size=sz, bold=bold, color=color)
def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center")
RGHT = Alignment(horizontal="right",  vertical="center")

AED      = '#,##0;(#,##0);"-"'
PCT      = '0.0%;(0.0%);"-"'
PCT2     = '0.00%;(0.00%);"-"'   # 2 decimal places — for editable scenario inputs
GP_DELTA = '[Color10]+#,##0;[Red]-#,##0;"-"'   # bright green positive / red negative
gl = get_column_letter


# ── XML-direct data loaders ───────────────────────────────────────────────────
def _parse_xlsx_rows(xlsx_bytes):
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        sheets = [n for n in zf.namelist() if 'worksheets/sheet' in n]
        with zf.open(sheets[0]) as f:
            content = f.read().decode('utf-8-sig')
    rows_xml = re.findall(r'<x:row>(.*?)</x:row>', content, re.DOTALL)
    result = []
    for row_xml in rows_xml:
        cells = re.findall(r'<x:c([^>]*)>(.*?)</x:c>', row_xml, re.DOTALL)
        row = []
        for attrs, inner in cells:
            inline = re.search(r'<x:is><x:t>(.*?)</x:t></x:is>', inner)
            v_tag  = re.search(r'<x:v>(.*?)</x:v>', inner)
            if inline:
                row.append(inline.group(1))
            elif v_tag:
                try:    row.append(float(v_tag.group(1)))
                except: row.append(None)
            else:
                row.append(None)
        result.append(row)
    return result


def _num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)):
        return float(v) if v == v else 0.0
    return 0.0


def load_d1(xlsx_bytes):
    all_rows = _parse_xlsx_rows(xlsx_bytes)

    # ── Pass 1: collect agency subtotals and raw per-agency customer TV sums ──
    agency_subtotals = {}
    agency_cu_raw    = {}
    cur_agency = None
    for row in all_rows[1:]:
        while len(row) < 20: row.append(None)
        ag, cu = row[0], row[1]
        if ag and isinstance(ag, str) and ag.strip():
            if ag.strip().lower() == 'total': continue
            cur_agency = ag.strip()
        if not cu or not cur_agency: continue
        cu_str = str(cu).strip() if isinstance(cu, str) else None
        if not cu_str: continue
        tv = _num(row[2]); gp = _num(row[9])
        if cu_str.lower() == 'total':
            agency_subtotals[cur_agency] = {'tv': tv, 'gp': gp}
        else:
            if cur_agency not in agency_cu_raw:
                agency_cu_raw[cur_agency] = {'tv': 0, 'gp': 0}
            agency_cu_raw[cur_agency]['tv'] += tv
            agency_cu_raw[cur_agency]['gp'] += gp

    # ── Pass 2: build customer list, scale TV and GP independently ─────────────
    # PowerBI individual rows don't always sum to agency subtotals.
    # We scale TV and GP by separate factors so both totals reconcile exactly.
    agencies  = {}
    customers = []
    cur_agency = None
    for row in all_rows[1:]:
        while len(row) < 20: row.append(None)
        ag, cu = row[0], row[1]
        if ag and isinstance(ag, str) and ag.strip():
            if ag.strip().lower() == 'total': continue
            cur_agency = ag.strip()
        if not cu or not cur_agency: continue
        cu_str = str(cu).strip() if isinstance(cu, str) else None
        if not cu_str or cu_str.lower() == 'total':
            if cu_str and cu_str.lower() == 'total':
                tv = _num(row[2]); gp = _num(row[9])
                agencies[cur_agency] = dict(agency=cur_agency, tv=tv, gp=gp)
            continue
        tv = _num(row[2]); gp = _num(row[9])
        if tv <= 0 and gp == 0: continue
        raw = agency_cu_raw.get(cur_agency, {'tv': 0, 'gp': 0})
        sub = agency_subtotals.get(cur_agency, {'tv': raw['tv'], 'gp': raw['gp']})
        scale_tv = (sub['tv'] / raw['tv']) if raw['tv'] else 1.0
        scale_gp = (sub['gp'] / raw['gp']) if raw['gp'] else 1.0
        customers.append(dict(
            agency=cur_agency, customer=cu_str,
            tv=tv * scale_tv, gp=gp * scale_gp
        ))

    ag_list = sorted(agencies.values(), key=lambda x: -x['tv'])
    cu_list = sorted(customers, key=lambda x: -x['tv'])
    return ag_list, cu_list


def load_d2(xlsx_bytes):
    all_rows = _parse_xlsx_rows(xlsx_bytes)

    # Pass 1: capture the authoritative Total row and raw sums
    total_tv = 0; total_gp = 0
    raw_tv   = 0; raw_gp   = 0
    for row in all_rows[1:]:
        while len(row) < 10: row.append(None)
        co = row[0]
        if not co or not isinstance(co, str): continue
        co = co.strip()
        tv = _num(row[1]); gp = _num(row[8])
        if co.lower() == 'total':
            total_tv = tv; total_gp = gp
        elif tv > 0 or gp != 0:
            raw_tv += tv; raw_gp += gp

    # Scale factors: bring row sums up to the Total row
    scale_tv = (total_tv / raw_tv) if raw_tv else 1.0
    scale_gp = (total_gp / raw_gp) if raw_gp else 1.0

    # Pass 2: build scaled destination rows
    rows = []
    for row in all_rows[1:]:
        while len(row) < 10: row.append(None)
        co = row[0]
        if not co or not isinstance(co, str): continue
        co = co.strip()
        if co.lower() == 'total': continue
        tv = _num(row[1]); gp = _num(row[8])
        if tv > 0 or gp != 0:
            rows.append(dict(country=co, tv=tv * scale_tv, gp=gp * scale_gp))
    return sorted(rows, key=lambda x: -x['tv'])


# ── Seasonality ───────────────────────────────────────────────────────────────
MONTHS  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
WEIGHTS = [0.16, 0.10, 0.02, 0.05, 0.05, 0.06, 0.06, 0.07, 0.10, 0.09, 0.10, 0.14]


def build_seasonality(ws, ag_rows, ytd_m):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ytd_tv = sum(r["tv"] for r in ag_rows)
    ytd_gp = sum(r["gp"] for r in ag_rows)
    ytd_wt = sum(WEIGHTS[:ytd_m])
    seas_row = 3 + ytd_m  # C{seas_row} = cumulative YTD factor, e.g. C6 for March

    # Row 1
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "SEASONALITY & EOY FORECAST  —  Booking Date 2026 YTD")
    c.font = hf(12); c.fill = fill(DARK_BLUE); c.alignment = CTR
    ws.row_dimensions[1].height = 26

    # Row 2
    for col, label, span, clr in [
        (1,"Month",1,MID_BLUE),(2,"2026 Weights",2,MID_BLUE),
        (4,"TV & GP  \u270e editable",2,"7F6000"),(6,"EOY Totals",2,GRN_HDR),(8,"Note",1,DARK_GREY)
    ]:
        if span > 1: ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+span-1)
        c = ws.cell(2, col, label); c.font = hf(9); c.fill = fill(clr); c.alignment = CTR
    ws.row_dimensions[2].height = 18

    # Row 3
    for ci, h in enumerate(["Month","Monthly Wt","Cumul.","TV (AED)","GP (AED)","EOY TV","EOY GP","Note"],1):
        c = ws.cell(3, ci, h); c.font = hf(9); c.fill = fill(MID_BLUE); c.alignment = CTR; c.border = bdr()
    ws.row_dimensions[3].height = 28

    for i, (mo, wt) in enumerate(zip(MONTHS, WEIGHTS)):
        r  = i + 4
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        completed = (i < ytd_m)
        ytd_row = 3 + ytd_m   # the row containing the last completed month's cumulative factor

        c = ws.cell(r, 1, mo); c.font = bf(bold=True); c.fill = fill(bg); c.alignment = LEFT; c.border = bdr()

        if completed:
            # Completed month weight — plain background, not editable highlight
            c = ws.cell(r, 2, wt); c.font = bf(bold=True, color="595959")
            c.fill = fill(bg); c.alignment = CTR; c.number_format = PCT; c.border = bdr()
        else:
            # Future month weight — gold, editable by user
            c = ws.cell(r, 2, wt); c.font = Font(name="Arial",size=9,color="0000FF")
            c.fill = fill(GOLD); c.alignment = CTR; c.number_format = PCT2; c.border = bdr()

        c = ws.cell(r, 3, f"=SUM($B$4:B{r})")
        c.font = bf(color="006100"); c.alignment = CTR; c.number_format = PCT; c.border = bdr()

        if completed:
            mo_tv = ytd_tv * (wt / ytd_wt) if ytd_wt else 0
            mo_gp = ytd_gp * (wt / ytd_wt) if ytd_wt else 0
            c = ws.cell(r, 4, round(mo_tv, 2)); c.font = Font(name="Arial",size=9,color="0000FF")
            c.fill = fill(GOLD); c.alignment = RGHT; c.number_format = AED; c.border = bdr()
            c = ws.cell(r, 5, round(mo_gp, 2)); c.font = Font(name="Arial",size=9,color="0000FF")
            c.fill = fill(GOLD); c.alignment = RGHT; c.number_format = AED; c.border = bdr()
            # FIXED: use SUM($D$4:$D$ytd_row)/$C$ytd_row*B{r}  NOT D{r}/C{r}*B{r}
            # Both completed AND future months must use the same annualisation base
            c = ws.cell(r, 6, f"=IFERROR(SUM($D$4:$D${ytd_row})/$C${ytd_row}*B{r},0)")
            c.font = bf(bold=True,color="006100"); c.fill = fill(LIGHT_GRN); c.alignment = RGHT; c.number_format = AED; c.border = bdr()
            c = ws.cell(r, 7, f"=IFERROR(SUM($E$4:$E${ytd_row})/$C${ytd_row}*B{r},0)")
            c.font = bf(bold=True,color="006100"); c.fill = fill(LIGHT_GRN); c.alignment = RGHT; c.number_format = AED; c.border = bdr()
            note = f"\u25cf  Actual \u00f7 YTD wt ({int(ytd_wt*100)}%) \u00d7 mo. wt"
        else:
            for col_n in [4, 5]:
                c = ws.cell(r, col_n, ""); c.fill = fill(bg); c.border = bdr()
            # Same formula as completed months — consistent annualisation
            c = ws.cell(r, 6, f"=IFERROR(SUM($D$4:$D${ytd_row})/$C${ytd_row}*B{r},0)")
            c.font = bf(color="006100"); c.fill = fill(LIGHT_GRN); c.alignment = RGHT; c.number_format = AED; c.border = bdr()
            c = ws.cell(r, 7, f"=IFERROR(SUM($E$4:$E${ytd_row})/$C${ytd_row}*B{r},0)")
            c.font = bf(color="006100"); c.fill = fill(LIGHT_GRN); c.alignment = RGHT; c.number_format = AED; c.border = bdr()
            note = f"\u25cb  Forecast: YTD \u00f7 {int(ytd_wt*100)}% \u00d7 {int(wt*100)}%"

        c = ws.cell(r, 8, note); c.font = bf(sz=8,color=DARK_GREY); c.alignment = LEFT; c.border = bdr()

    # Row 16 TOTAL — C16 = live YTD factor used by all analysis sheets
    tr = 16
    c = ws.cell(tr, 1, "TOTAL"); c.font = hf(9); c.fill = fill(DARK_BLUE); c.alignment = LEFT; c.border = bdr()
    c = ws.cell(tr, 2, "=SUM(B4:B15)"); c.font = hf(9); c.fill = fill(DARK_BLUE); c.alignment = CTR; c.number_format = PCT; c.border = bdr()
    c = ws.cell(tr, 3, f"=C{seas_row}")  # C16 = live cumulative YTD factor = SUM(B4:B{seas_row-1})
    # This means: change any weight in col B → C column recalculates → C16 updates → all analysis tab EOY forecasts update
    c.font = hf(9,bold=True,color="FFFF00"); c.fill = fill(DARK_BLUE); c.alignment = CTR; c.number_format = PCT; c.border = bdr()
    for ci, fmt in [(4,AED),(5,AED),(6,AED),(7,AED)]:
        c = ws.cell(tr, ci, f"=SUM({gl(ci)}4:{gl(ci)}15)")
        c.font = hf(9); c.fill = fill(DARK_BLUE); c.alignment = RGHT; c.number_format = fmt; c.border = bdr()
    ws.cell(tr, 8, "").border = bdr()

    # Row 18 note
    eoy_tv = ytd_tv / ytd_wt if ytd_wt else 0
    eoy_gp = ytd_gp / ytd_wt if ytd_wt else 0
    ws.merge_cells("A18:H18")
    c = ws.cell(18, 1,
        f"YTD ({'-'.join(MONTHS[:ytd_m])}): {ytd_tv:,.0f} AED TV  |  "
        f"YTD weight: {int(ytd_wt*100)}%  ->  EOY projection: {eoy_tv:,.0f} AED TV  /  {eoy_gp:,.0f} AED GP.  "
        f"Each month = (YTD total / YTD weight) x month weight.  Yellow cells are editable.  "
        f"C16 = YTD cum. weight ({int(ytd_wt*100)}%) used in all other tabs.")
    c.font = bf(sz=8,color=DARK_GREY); c.alignment = LEFT

    for ci, w in enumerate([10,13,10,16,16,16,16,42],1):
        ws.column_dimensions[gl(ci)].width = w


def build_analysis_sheet(ws, title, rows, id_key, id_label, agency_key=None, master_tv=None):
    """
    Column layout (after n_id identity cols):
    YTD ACTUALS (3):    TV | GP | GP%
    SCENARIO INPUTS (2): GP%Adj | TV_Chg%
    EOY BASE FCST (3):  Base_EOY_TV | Base_EOY_GP | Base_EOY_GP%
                        (pure seasonality projection, no adjustments)
    ADJUSTED FCST (3):  Adj_EOY_TV | Adj_EOY_GP | GP_Delta
                        (includes TV_Chg% and GP%_Adj)
    Total data cols = 11, N_COLS = n_id + 11
    """
    ws.sheet_view.showGridLines = False

    n_id       = 2 if agency_key else 1
    N_COLS     = n_id + 11
    DATA_START = 8
    n_rows     = len(rows)
    DATA_END   = DATA_START + n_rows - 1

    # Column positions (1-based)
    tv_col    = n_id + 1   # YTD TV
    gp_col    = n_id + 2   # YTD GP
    gpp_col   = n_id + 3   # YTD GP%
    adj_col   = n_id + 4   # GP% Adj (editable)
    tvc_col   = n_id + 5   # TV Chg% (editable)
    beov_col  = n_id + 6   # Base EOY TV  (no TV_Chg)
    beog_col  = n_id + 7   # Base EOY GP  (= Base_EOY_TV × GP%)
    beopc_col = n_id + 8   # Base EOY GP% (= GP% always)
    aeov_col  = n_id + 9   # Adj EOY TV   (Base × (1+TV_Chg%))
    aeog_col  = n_id + 10  # Adj EOY GP   (Adj_TV × (GP%+GP%Adj))
    delta_col = n_id + 11  # GP Delta     (Adj_GP − Base_GP)

    tv_l   = gl(tv_col);  gp_l   = gl(gp_col);  gpp_l  = gl(gpp_col)
    adj_l  = gl(adj_col); tvc_l  = gl(tvc_col)
    beov_l = gl(beov_col); beog_l = gl(beog_col); beopc_l = gl(beopc_col)
    aeov_l = gl(aeov_col); aeog_l = gl(aeog_col)
    SEAS   = "Seasonality!$F$16"  # full-year EOY TV total from Seasonality tab

    # ── KPI banner (rows 1-2) ────────────────────────────────────────────────
    kpis = [
        ("YTD Total Value",   f"=SUM({tv_l}{DATA_START}:{tv_l}{DATA_END})",       AED, MID_BLUE,  "1F3864", 2),
        ("YTD Gross Profit",  f"=SUM({gp_l}{DATA_START}:{gp_l}{DATA_END})",       AED, MID_BLUE,  "1F3864", 2),
        ("YTD GP%",           f"=IFERROR(SUM({gp_l}{DATA_START}:{gp_l}{DATA_END})/IF(SUM({tv_l}{DATA_START}:{tv_l}{DATA_END})=0,1,SUM({tv_l}{DATA_START}:{tv_l}{DATA_END})),0)", PCT, MID_BLUE, "1F3864", 2),
        ("EOY TV (Base)",     f"=SUM({beov_l}{DATA_START}:{beov_l}{DATA_END})",    AED, GRN_HDR,  "375623", 2),
        ("EOY GP (Base)",     f"=SUM({beog_l}{DATA_START}:{beog_l}{DATA_END})",    AED, GRN_HDR,  "375623", 2),
        ("EOY GP (Adjusted)", f"=SUM({aeog_l}{DATA_START}:{aeog_l}{DATA_END})",    AED, "7F3F00", "7F3F00", 2),
    ]
    col = 1
    for label, formula, fmt, bg_clr, val_clr, span in kpis:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+span-1)
        c = ws.cell(1, col, label)
        c.font = hf(9); c.fill = fill(bg_clr); c.alignment = CTR; c.border = bdr()
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+span-1)
        c = ws.cell(2, col, formula)
        c.font = Font(name="Arial", size=11, bold=True, color=val_clr)
        c.fill = fill(WHITE); c.alignment = RGHT; c.number_format = fmt; c.border = bdr()
        col += span
    ws.row_dimensions[1].height = 20; ws.row_dimensions[2].height = 26
    for r in [3,4]: ws.row_dimensions[r].height = 6

    # ── Row 5: title ─────────────────────────────────────────────────────────
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=N_COLS)
    c = ws.cell(5, 1, title); c.font = hf(13); c.fill = fill(DARK_BLUE); c.alignment = CTR
    ws.row_dimensions[5].height = 28

    # ── Row 6: colour bands ──────────────────────────────────────────────────
    # n_id | 3 YTD ACTUALS | 2 SCENARIO INPUTS | 3 EOY BASE FCST | 3 ADJUSTED FCST
    bands = [(n_id,"",DARK_BLUE),(3,"◀  YTD ACTUALS",MID_BLUE),
             (2,"✏  SCENARIO INPUTS","7F3F00"),(3,"▶  EOY BASE FCST",GRN_HDR),
             (3,"★  ADJUSTED FCST","375E23")]
    col = 1
    for span, label, clr in bands:
        if span > 1: ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+span-1)
        c = ws.cell(6, col, label); c.font = hf(9); c.fill = fill(clr); c.alignment = CTR; c.border = bdr()
        col += span
    ws.row_dimensions[6].height = 18

    # ── Row 7: column headers ─────────────────────────────────────────────────
    if agency_key:
        hdrs = [" Agency", id_label,
                "YTD TV (AED)", "YTD GP (AED)", "YTD GP%",
                "GP% Adj (+pp)", "TV Chg %",
                "EOY TV", "EOY GP", "EOY GP%",
                "Adj. EOY TV", "Adj. EOY GP", "GP Delta"]
    else:
        hdrs = [id_label,
                "YTD TV (AED)", "YTD GP (AED)", "YTD GP%",
                "GP% Adj (+pp)", "TV Chg %",
                "EOY TV", "EOY GP", "EOY GP%",
                "Adj. EOY TV", "Adj. EOY GP", "GP Delta"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(7, ci, h); c.font = hf(9); c.fill = fill(MID_BLUE); c.alignment = CTR; c.border = bdr()
    ws.row_dimensions[7].height = 32
    ws.freeze_panes = "A8"

    # ── Rows 8+: data ─────────────────────────────────────────────────────────
    TV_SUM = f"SUM({tv_l}{DATA_START}:{tv_l}{DATA_END})"
    # Denominator for proportional EOY share — always the master agency total so all
    # tabs produce consistent EOY TV values for the same underlying booking volume.
    # Falls back to the tab's own SUM if master_tv not provided.
    TV_DENOM = str(round(master_tv)) if master_tv else TV_SUM

    for ri, row in enumerate(rows):
        r   = DATA_START + ri
        is_tot = row.get("is_total", False)
        bg  = "D9E1F2" if is_tot else (PALE_BLUE if ri % 2 == 0 else WHITE)

        # Identity cols
        col = 1
        if agency_key:
            c = ws.cell(r, col, row.get(agency_key, ""))
            c.font = bf(bold=is_tot); c.alignment = LEFT; c.fill = fill(bg); c.border = bdr(); col += 1
        c = ws.cell(r, col, row.get(id_key, ""))
        c.font = bf(bold=True); c.alignment = LEFT; c.fill = fill(bg); c.border = bdr()

        # YTD TV
        c = ws.cell(r, tv_col, row.get("tv", 0) or 0)
        c.font = bf(); c.alignment = RGHT; c.fill = fill(bg); c.border = bdr(); c.number_format = AED
        # YTD GP
        c = ws.cell(r, gp_col, row.get("gp", 0) or 0)
        c.font = bf(); c.alignment = RGHT; c.fill = fill(bg); c.border = bdr(); c.number_format = AED
        # YTD GP%
        c = ws.cell(r, gpp_col, f"=IFERROR({gp_l}{r}/IF({tv_l}{r}=0,1,{tv_l}{r}),0)")
        c.font = bf(); c.alignment = CTR; c.fill = fill(bg); c.border = bdr(); c.number_format = PCT

        # GP% Adj (editable, gold, 2dp)
        c = ws.cell(r, adj_col, 0)
        c.font = Font(name="Arial",size=9,color="0000FF")
        c.alignment = CTR; c.fill = fill(GOLD); c.border = bdr(); c.number_format = PCT2
        # TV Chg% (editable, gold, 2dp)
        c = ws.cell(r, tvc_col, 0)
        c.font = Font(name="Arial",size=9,color="0000FF")
        c.alignment = CTR; c.fill = fill(GOLD); c.border = bdr(); c.number_format = PCT2

        # Base EOY TV = proportional share of Seasonality total (NO TV_Chg%)
        # Uses master agency TV as denominator — consistent across all tabs
        c = ws.cell(r, beov_col,
            f"=IFERROR({tv_l}{r}/IF({TV_DENOM}=0,1,{TV_DENOM})*{SEAS},0)")
        c.font = bf(bold=True,color="006100"); c.alignment = RGHT
        c.fill = fill(LIGHT_GRN); c.border = bdr(); c.number_format = AED

        # Base EOY GP = Base_EOY_TV × GP%  → EOY GP% = YTD GP% always
        c = ws.cell(r, beog_col,
            f"=IFERROR({beov_l}{r}*{gpp_l}{r},0)")
        c.font = bf(bold=True,color="006100"); c.alignment = RGHT
        c.fill = fill(LIGHT_GRN); c.border = bdr(); c.number_format = AED

        # Base EOY GP% = always equals YTD GP%
        c = ws.cell(r, beopc_col,
            f"=IFERROR({beog_l}{r}/IF({beov_l}{r}=0,1,{beov_l}{r}),0)")
        c.font = bf(color="006100"); c.alignment = CTR
        c.fill = fill(LIGHT_GRN); c.border = bdr(); c.number_format = PCT

        # Adj EOY TV = Base_EOY_TV × (1 + TV_Chg%)
        c = ws.cell(r, aeov_col,
            f"=IFERROR({beov_l}{r}*(1+{tvc_l}{r}),0)")
        c.font = bf(bold=True); c.alignment = RGHT
        c.fill = fill(LIGHT_BLUE); c.border = bdr(); c.number_format = AED

        # Adj EOY GP = Adj_EOY_TV × (GP% + GP%_Adj)
        c = ws.cell(r, aeog_col,
            f"=IFERROR({aeov_l}{r}*({gpp_l}{r}+{adj_l}{r}),0)")
        c.font = bf(bold=True); c.alignment = RGHT
        c.fill = fill(LIGHT_BLUE); c.border = bdr(); c.number_format = AED

        # GP Delta = Adj_GP − Base_GP  (captures volume + margin)
        c = ws.cell(r, delta_col,
            f"=IFERROR({aeog_l}{r}-{beog_l}{r},0)")
        c.font = bf(); c.alignment = RGHT
        c.fill = fill(LIGHT_BLUE); c.border = bdr(); c.number_format = GP_DELTA

    # ── TOTAL row ─────────────────────────────────────────────────────────────
    tr = DATA_END + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=n_id)
    c = ws.cell(tr, 1, "TOTAL"); c.font = hf(9); c.fill = fill(DARK_BLUE); c.alignment = LEFT; c.border = bdr()

    for ci in range(tv_col, delta_col+1):
        c = ws.cell(tr, ci)
        if ci in {adj_col, tvc_col}:
            c.value = ""
        elif ci in {gpp_col, beopc_col}:
            c.value = f"=IFERROR(SUM({gp_l}{DATA_START}:{gp_l}{DATA_END})/IF(SUM({tv_l}{DATA_START}:{tv_l}{DATA_END})=0,1,SUM({tv_l}{DATA_START}:{tv_l}{DATA_END})),0)"
            c.number_format = PCT; c.alignment = CTR
        else:
            c.value = f"=SUM({gl(ci)}{DATA_START}:{gl(ci)}{DATA_END})"
            c.number_format = AED if ci != delta_col else GP_DELTA
            c.alignment = RGHT
        c.font = hf(9); c.fill = fill(DARK_BLUE); c.border = bdr()
    ws.row_dimensions[tr].height = 18

    # ── Notes ─────────────────────────────────────────────────────────────────
    nr = tr + 1
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=N_COLS)
    c = ws.cell(nr, 1,
        "EOY Base = seasonality projection at current GP%, no adjustments.  "
        "TV Chg% grows Adj EOY TV vs Base EOY TV.  "
        "GP% Adj (+pp) adds margin points: e.g. GP%=2.8%, Adj=0.20% \u2192 adjusted=3.0%.  "
        "GP Delta = Adj GP \u2212 Base GP (captures both volume growth and margin improvement).")
    c.font = bf(sz=8,color=DARK_GREY); c.alignment = LEFT
    ws.row_dimensions[nr].height = 14

    # ── Opportunity Insights (tab-specific, purple header) ───────────────────
    tv_total   = sum(r["tv"] for r in rows)
    gp_total   = sum(r["gp"] for r in rows)
    avg_gp_pct = gp_total / tv_total if tv_total else 0

    PANEL_A_CLR = "9B2DC9"   # purple  – Improve Margin
    PANEL_B_CLR = "C00000"   # dark red – Grow Volume
    PANEL_C_CLR = "375623"   # dark green – Biggest GP Gap

    ir = nr + 2
    ws.merge_cells(start_row=ir, start_column=1, end_row=ir, end_column=N_COLS)
    c = ws.cell(ir, 1,
        f"  \u2605  OPPORTUNITY INSIGHTS  \u2014  Where managers can act  "
        f"|  Avg GP% for this view: {avg_gp_pct*100:.1f}%")
    c.font = hf(10); c.fill = fill(INSIGHT_BG); c.alignment = LEFT
    ws.row_dimensions[ir].height = 20

    # Panel headers — 3 equal panels, each exactly PANEL_W columns wide
    # Use floor division to get equal thirds, give remainder to last panel
    PANEL_W = N_COLS // 3
    p1_end  = PANEL_W
    p2_end  = PANEL_W * 2
    p3_end  = N_COLS   # last panel gets any remainder too

    ir2 = ir + 1
    desc_data = [
        (1,      p1_end, "A \u2014 Improve Margin", PANEL_A_CLR,
         "Large TV but below-avg GP%. Raising GP% to avg = biggest absolute gain."),
        (p1_end+1, p2_end, "B \u2014 Grow Volume",    PANEL_B_CLR,
         "Above-avg GP%, low volume. Growing TV = high ROI per AED spent."),
        (p2_end+1, p3_end, "C \u2014 Biggest GP Gap", PANEL_C_CLR,
         "Largest absolute gap between current GP and what the avg margin would generate."),
    ]
    for c1, c2, hdr, clr, desc in desc_data:
        ws.merge_cells(start_row=ir2,   start_column=c1, end_row=ir2,   end_column=c2)
        c = ws.cell(ir2, c1, hdr); c.font = hf(9); c.fill = fill(clr); c.alignment = CTR; c.border = bdr()
        ws.merge_cells(start_row=ir2+1, start_column=c1, end_row=ir2+1, end_column=c2)
        c = ws.cell(ir2+1, c1, desc); c.font = bf(sz=8,color=DARK_GREY); c.alignment = LEFT
    ws.row_dimensions[ir2].height = 18; ws.row_dimensions[ir2+1].height = 14

    ir3 = ir2 + 2
    # Column headers: match panel column boundaries exactly
    panel_hdrs = [
        (1,      p1_end,  ["Name","YTD TV","GP%","Opportunity"][:p1_end],  PANEL_A_CLR),
        (p1_end+1, p2_end, ["Name","YTD TV","GP%","Opportunity"][:p2_end-p1_end], PANEL_B_CLR),
        (p2_end+1, p3_end, ["Name","YTD TV","GP%","Opportunity"][:p3_end-p2_end], PANEL_C_CLR),
    ]
    for c_start, c_end, hdrs_list, clr in panel_hdrs:
        for idx, h in enumerate(hdrs_list):
            ci = c_start + idx
            if ci <= N_COLS:
                c = ws.cell(ir3, ci, h)
                c.font = hf(9); c.fill = fill(clr); c.alignment = CTR; c.border = bdr()
    ws.row_dimensions[ir3].height = 18

    data_rows_with_tv = [r for r in rows if r["tv"] > 0]
    panel_a = sorted([r for r in data_rows_with_tv if (r["gp"]/r["tv"]) < avg_gp_pct],
                     key=lambda x: -(x["tv"] * (avg_gp_pct - x["gp"]/x["tv"])))[:5]
    panel_b = sorted([r for r in data_rows_with_tv if (r["gp"]/r["tv"]) > avg_gp_pct],
                     key=lambda x: x["tv"])[:5]
    panel_c = sorted(data_rows_with_tv,
                     key=lambda x: -(abs(x["tv"] * avg_gp_pct - x["gp"])))[:5]

    # Column offsets within each panel (0-based within panel start)
    for pi in range(max(len(panel_a), len(panel_b), len(panel_c))):
        pr = ir3 + 1 + pi
        bg = LIGHT_GREY if pi % 2 == 0 else WHITE

        # Panel A cols: 1 .. p1_end
        if pi < len(panel_a):
            ra = panel_a[pi]; ra_pct = ra["gp"]/ra["tv"]
            upside = ra["tv"] * (avg_gp_pct - ra_pct)
            pa_vals = [(ra[id_key],None),(ra["tv"],AED),(ra_pct,PCT),(f"{upside:,.0f} AED upside",None)]
            for idx, (val, fmt) in enumerate(pa_vals[:p1_end]):
                ci = 1 + idx
                c = ws.cell(pr,ci,val); c.font=bf(sz=8); c.fill=fill(bg); c.border=bdr()
                if fmt: c.number_format=fmt

        # Panel B cols: p1_end+1 .. p2_end
        if pi < len(panel_b):
            rb = panel_b[pi]; rb_pct = rb["gp"]/rb["tv"]
            pb_vals = [(rb[id_key],None),(rb["tv"],AED),(rb_pct,PCT),(f"High margin",None)]
            for idx, (val, fmt) in enumerate(pb_vals[:p2_end-p1_end]):
                ci = p1_end + 1 + idx
                if ci <= N_COLS:
                    c = ws.cell(pr,ci,val); c.font=bf(sz=8); c.fill=fill(bg); c.border=bdr()
                    if fmt: c.number_format=fmt

        # Panel C cols: p2_end+1 .. p3_end
        if pi < len(panel_c):
            rc = panel_c[pi]; rc_pct = rc["gp"]/rc["tv"]
            gap = rc["tv"] * avg_gp_pct - rc["gp"]
            pc_vals = [(rc[id_key],None),(rc["tv"],AED),(rc_pct,PCT),(f"{gap:,.0f} AED {'gap' if gap>0 else 'above avg'}",None)]
            for idx, (val, fmt) in enumerate(pc_vals[:p3_end-p2_end]):
                ci = p2_end + 1 + idx
                if ci <= N_COLS:
                    c = ws.cell(pr,ci,val); c.font=bf(sz=8); c.fill=fill(bg); c.border=bdr()
                    if fmt: c.number_format=fmt

        ws.row_dimensions[pr].height = 14

    # ── Operational Cadence ───────────────────────────────────────────────────
    ops_start = ir3 + 1 + max(len(panel_a), len(panel_b), len(panel_c)) + 2

    ws.merge_cells(start_row=ops_start, start_column=1, end_row=ops_start, end_column=N_COLS)
    c = ws.cell(ops_start, 1, "  \u23f1  OPERATIONAL CADENCE  \u2014  Recommended actions by frequency")
    c.font = hf(10); c.fill = fill(INSIGHT_BG); c.alignment = LEFT
    ws.row_dimensions[ops_start].height = 20

    ops_sections = [
        ("DAILY", OPS_DAILY, [
            "Check prior-day bookings vs daily run-rate target  (Total TV \u00f7 working days remaining)",
            "Flag any agency / customer below 80% of their daily pace",
            "Review new GP% on confirmed bookings — escalate if below floor margin",
            "Update GP% Adj inputs for any accounts with fresh pricing changes",
        ]),
        ("WEEKLY", OPS_WEEKLY, [
            "Compare YTD TV and GP vs prior week — calculate weekly growth rate",
            "Run Opportunity Panel A: identify accounts drifting below avg GP% for margin coaching",
            "Run Opportunity Panel B: contact high-margin low-volume accounts with volume incentives",
            "Review Seasonality weights — adjust monthly targets if booking patterns shift",
            "Export updated modeller from Drive and share with team / stakeholders",
        ]),
        ("MONTHLY", OPS_MONTHLY, [
            "Full reconciliation: YTD actual vs EOY forecast — update TV Chg% inputs",
            "Refresh data_1.xlsx and data_2.xlsx from PowerBI and drop into Drive folder",
            "Review GP Delta column — accounts with negative delta need action plans",
            "Update seasonality completed months (actual TV/GP in Seasonality tab)",
            "Run Panel C: GP Gap analysis — build targeted recovery plans for top 5 accounts",
            "Board / management report: pull Dashboard KPIs for monthly review pack",
        ]),
    ]

    r_ops = ops_start + 1
    for period, clr, actions in ops_sections:
        ws.merge_cells(start_row=r_ops, start_column=1, end_row=r_ops, end_column=N_COLS)
        c = ws.cell(r_ops, 1, f"  {period}")
        c.font = bf(sz=9,bold=True,color="000000"); c.fill = fill(clr); c.alignment = LEFT; c.border = bdr()
        ws.row_dimensions[r_ops].height = 16
        r_ops += 1
        for action in actions:
            ws.merge_cells(start_row=r_ops, start_column=1, end_row=r_ops, end_column=N_COLS)
            c = ws.cell(r_ops, 1, f"    \u2022  {action}")
            c.font = bf(sz=8); c.fill = fill(clr); c.alignment = LEFT
            ws.row_dimensions[r_ops].height = 14
            r_ops += 1
        r_ops += 1  # spacer between sections

    # Column widths
    widths = [16,28,14,14,9,9,9,16,14,9,14,14] if agency_key else [28,14,14,9,9,9,16,14,9,14,14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[gl(i)].width = w


def build_dashboard(ws, today_str, data_month, ag_rows, de_rows, ytd_wt):
    ws.sheet_view.showGridLines = False

    ytd_tv = sum(r["tv"] for r in ag_rows)
    ytd_gp = sum(r["gp"] for r in ag_rows)
    gp_pct = ytd_gp / ytd_tv if ytd_tv else 0
    eoy_tv = ytd_tv / ytd_wt if ytd_wt else 0
    eoy_gp = ytd_gp / ytd_wt if ytd_wt else 0

    # Row 1: title
    ws.merge_cells("A1:P1")
    c = ws.cell(1, 1,
        f"ELEVATE DMC \u2014 PROFITABILITY MODELLER 2026  |  Refreshed: {today_str}  |  "
        f"Booking Data through: {data_month}")
    c.font = hf(13); c.fill = fill(DARK_BLUE); c.alignment = CTR
    ws.row_dimensions[1].height = 30

    ws.row_dimensions[2].height = 8

    # Rows 3-4: KPI banner — label row coloured bg/white text, value row white bg/coloured text
    # EOY TV/GP reference Seasonality!F16/G16 so changing weights updates the dashboard live
    kpi_specs = [
        ("YTD Total Value",   ytd_tv,                 AED, MID_BLUE,  "1F3864", False),
        ("YTD Gross Profit",  ytd_gp,                 AED, MID_BLUE,  "1F3864", False),
        ("YTD GP%",           gp_pct,                 PCT, MID_BLUE,  "1F3864", False),
        ("EOY TV Forecast",   "=Seasonality!$F$16",   AED, GRN_HDR,   "375623", True),
        ("EOY GP Forecast",   "=Seasonality!$G$16",   AED, GRN_HDR,   "375623", True),
        ("EOY GP% Forecast",  gp_pct,                 PCT, GRN_HDR,   "375623", False),
    ]
    col = 1
    for label, val, fmt, bg_clr, val_clr, is_formula in kpi_specs:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)
        c = ws.cell(3, col, label); c.font = hf(9); c.fill = fill(bg_clr); c.alignment = CTR; c.border = bdr()
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+2)
        c = ws.cell(4, col, val)
        c.font = Font(name="Arial", size=12, bold=True, color=val_clr)
        c.fill = fill(WHITE); c.alignment = RGHT; c.number_format = fmt; c.border = bdr()
        col += 3
    ws.row_dimensions[3].height = 20; ws.row_dimensions[4].height = 28

    for r in [5,6,8,9,10]: ws.row_dimensions[r].height = 6

    # Row 7: how-to
    ws.merge_cells("A7:P7")
    c = ws.cell(7, 1,
        "HOW TO USE: Edit TV/GP in Agency Groups / Customer / Destination tabs \u2192 KPI banner and EOY update live.  "
        "GP% Adj (+pp) adds percentage POINTS to current GP%: e.g. GP%=2.8%, Adj=0.2% \u2192 adjusted=3.0%.  "
        "Seasonality tab weights drive EOY projections (yellow cells editable).")
    c.font = bf(sz=8,color=DARK_GREY); c.fill = fill(LIGHT_GREY); c.alignment = LEFT
    ws.row_dimensions[7].height = 14

    # Top agencies (rows 11-20+)
    ws.merge_cells("A11:G11")
    c = ws.cell(11, 1, "TOP AGENCY GROUPS \u2014 YTD GROSS PROFIT")
    c.font = hf(10); c.fill = fill(DARK_BLUE); c.alignment = CTR
    ws.row_dimensions[11].height = 20

    for ci, h in enumerate(["Agency Group","YTD TV","YTD GP","GP%","EOY TV","EOY GP","EOY GP%"],1):
        c = ws.cell(12, ci, h); c.font = hf(9); c.fill = fill(MID_BLUE); c.alignment = CTR; c.border = bdr()
    ws.row_dimensions[12].height = 22

    for ri, row in enumerate(sorted(ag_rows, key=lambda x: -x["tv"])):
        r   = 13 + ri
        bg  = PALE_BLUE if ri % 2 == 0 else WHITE
        gpp = row["gp"] / row["tv"] if row["tv"] else 0
        etv = row["tv"] / ytd_wt if ytd_wt else 0
        egp = row["gp"] / ytd_wt if ytd_wt else 0
        for ci,(val,fmt) in enumerate([(row["agency"],None),(row["tv"],AED),(row["gp"],AED),(gpp,PCT),(etv,AED),(egp,AED),(gpp,PCT)],1):
            c = ws.cell(r,ci,val); c.font=bf(); c.fill=fill(bg); c.border=bdr()
            if fmt: c.number_format=fmt
            c.alignment=RGHT if fmt else LEFT
        ws.row_dimensions[r].height = 16

    # Top destinations (cols H-N)
    ws.merge_cells("H11:N11")
    c = ws.cell(11, 8, "TOP DESTINATIONS \u2014 YTD GROSS PROFIT")
    c.font = hf(10); c.fill = fill(DARK_BLUE); c.alignment = CTR

    for ci, h in enumerate(["Destination","YTD TV","YTD GP","GP%","EOY TV","EOY GP","EOY GP%"],1):
        c = ws.cell(12, 7+ci, h); c.font = hf(9); c.fill = fill(MID_BLUE); c.alignment = CTR; c.border = bdr()

    for ri, row in enumerate(sorted(de_rows, key=lambda x: -x["tv"])[:10]):
        r   = 13 + ri
        bg  = PALE_BLUE if ri % 2 == 0 else WHITE
        gpp = row["gp"] / row["tv"] if row["tv"] else 0
        etv = row["tv"] / ytd_wt if ytd_wt else 0
        egp = row["gp"] / ytd_wt if ytd_wt else 0
        for ci,(val,fmt) in enumerate([(row["country"],None),(row["tv"],AED),(row["gp"],AED),(gpp,PCT),(etv,AED),(egp,AED),(gpp,PCT)],1):
            c = ws.cell(r,7+ci,val); c.font=bf(); c.fill=fill(bg); c.border=bdr()
            if fmt: c.number_format=fmt
            c.alignment=RGHT if fmt else LEFT
        ws.row_dimensions[r].height = 16

    # Column widths
    for ci, w in enumerate([22,16,16,9,16,16,9,4,22,16,16,9,16,16,9],1):
        ws.column_dimensions[gl(ci)].width = w


# ── main rebuild ──────────────────────────────────────────────────────────────
def rebuild(d1_bytes, d2_bytes):
    ag_rows, cu_rows = load_d1(d1_bytes)
    de_rows          = load_d2(d2_bytes)

    today      = datetime.date.today()
    today_str  = today.strftime("%d %b %Y")
    ytd_m      = today.month
    ytd_wt     = sum(WEIGHTS[:ytd_m])
    data_month = MONTHS[ytd_m - 1] + " 2026"

    wb = openpyxl.Workbook()
    ss = wb.active;              ss.title = "Seasonality"   # ← first tab
    db = wb.create_sheet("Dashboard")
    ag = wb.create_sheet("Agency Groups")
    cu = wb.create_sheet("Customer")
    de = wb.create_sheet("Destination")

    # master_tv for agency + customer tabs = agency subtotal total (authoritative)
    master_tv    = sum(r["tv"] for r in ag_rows)
    # master_tv for destination = destination's own scaled total (different export)
    master_tv_de = sum(r["tv"] for r in de_rows)

    build_seasonality(ss, ag_rows, ytd_m)
    build_dashboard(db, today_str, data_month, ag_rows, de_rows, ytd_wt)
    build_analysis_sheet(ag, "AGENCY GROUPS ANALYSIS", ag_rows, id_key="agency",   id_label="Agency Group", master_tv=master_tv)
    build_analysis_sheet(cu, "CUSTOMER ANALYSIS",      cu_rows, id_key="customer", id_label="Customer", agency_key="agency", master_tv=master_tv)
    build_analysis_sheet(de, "DESTINATION ANALYSIS",   de_rows, id_key="country",  id_label="Country",  master_tv=master_tv_de)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


# ── Flask routes ──────────────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "profitability-modeller"})


@app.route("/rebuild", methods=["POST"])
def rebuild_endpoint():
    try:
        body     = request.get_json()
        d1_bytes = base64.b64decode(body["data1_b64"])
        d2_bytes = base64.b64decode(body["data2_b64"])
        result   = rebuild(d1_bytes, d2_bytes)
        return jsonify({"modeller_b64": base64.b64encode(result).decode(), "status": "ok"})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/rebuild-from-drive", methods=["POST"])
def rebuild_from_drive():
    import requests as req
    try:
        body        = request.get_json()
        token       = body["access_token"]
        folder_id   = body["folder_id"]
        output_name = body.get("output_name", "Profitability_Modeller_2026.xlsx")
        headers     = {"Authorization": f"Bearer {token}"}

        def find_file(name):
            q = f"'{folder_id}' in parents and name='{name}' and trashed=false"
            r = req.get("https://www.googleapis.com/drive/v3/files",
                        params={"q": q, "fields": "files(id,name)", "pageSize": 1},
                        headers=headers, timeout=30)
            r.raise_for_status()
            files = r.json().get("files", [])
            if not files: raise FileNotFoundError(f"{name} not found in Drive folder")
            return files[0]["id"]

        def download_file(file_id):
            r = req.get(f"https://www.googleapis.com/drive/v3/files/{file_id}",
                        params={"alt": "media"}, headers=headers, timeout=60)
            r.raise_for_status()
            return r.content

        d1     = download_file(find_file("data_1.xlsx"))
        d2     = download_file(find_file("data_2.xlsx"))
        result = rebuild(d1, d2)

        mime  = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        q_out = f"'{folder_id}' in parents and name='{output_name}' and trashed=false"
        r_out = req.get("https://www.googleapis.com/drive/v3/files",
                        params={"q": q_out, "fields": "files(id)", "pageSize": 1},
                        headers=headers, timeout=30)
        r_out.raise_for_status()
        existing = r_out.json().get("files", [])

        if existing:
            up = req.patch(
                f"https://www.googleapis.com/upload/drive/v3/files/{existing[0]['id']}",
                params={"uploadType": "media"},
                headers={**headers, "Content-Type": mime},
                data=result, timeout=120)
        else:
            import json as _json
            meta = _json.dumps({"name": output_name, "parents": [folder_id]})
            up = req.post("https://www.googleapis.com/upload/drive/v3/files",
                          params={"uploadType": "multipart"}, headers=headers,
                          files={"metadata": ("meta", meta, "application/json"),
                                 "file": ("file", result, mime)}, timeout=120)
        up.raise_for_status()
        return jsonify({"status": "ok", "output_file": output_name,
                        "refreshed": datetime.date.today().strftime("%d %b %Y")})

    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
